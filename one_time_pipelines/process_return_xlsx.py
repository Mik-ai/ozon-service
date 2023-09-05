import openpyxl, re
import aiohttp, asyncio
import json

CRM_API_KEY = "FQyXv6xsDM2fAHlST7Dui3zErYIPuXaR"


def dict_batches_100(data: dict) -> list:
    """function for slising dict to batches, is used cause CRM could process up to 100 orders per request

    Args:
        data (dict): dict of orders

    Returns:
        list: list of dicts
    """
    key_batches = [[key for key in data][_ : _ + 100] for _ in range(0, len(data), 100)]
    result = [dict([(key, data[key]) for key in batch]) for batch in key_batches]
    return result


def extract_data_wb(wb: openpyxl.Workbook) -> dict:
    """extracting result from wildberries excel table, does not care about duplicates

    Args:
        wb (openpyxl.Workbook): workbook

    Returns:
        dict: dict of orders
    """
    columns = (0, 14)
    sh = wb.active
    data = [[str(x[m].value) for m in columns] for x in sh]
    result = {}

    for i in data:
        if i[1] != "Отказ покупателем":
            continue
        x = {}
        x["status"] = "vozvrat-poluchen-wb-fbo"
        x["comment"] = ""
        result[i[0]] = x
    return result


def extract_data_ya(wb: openpyxl.Workbook) -> dict:
    """extracting result from yandex excel table, does not care about duplicates

    Args:
        wb (openpyxl.Workbook): workbook

    Returns:
        dict: dict of orders
    """
    columns = (7, 21, 24)
    sh = wb.active
    data = [[str(x[m].value) for m in columns] for x in sh]
    result = {}

    for i in data:
        if not i[1] in ["Выдан", "Отменён"] or bool(
            re.match(pattern="[A-zА-я]", string=str(i[0]), flags=re.IGNORECASE)
        ):
            continue
        if x := result.get(i[0], {}):
            result.pop(i[0])
        else:
            i[0] = re.sub("\.0", "", i[0])
            x = {}
            x["status"] = "vozvrat-predvaritelno-poluchen"
            x["comment"] = i[2]
            result[i[0]] = x
    return result


def extract_data_ozon(wb: openpyxl.Workbook) -> dict:
    """extracting result from ozon excel table, will find and process duplicates

    Args:
        wb (openpyxl.Workbook): workbook

    Returns:
        dict: dict of orders
    """
    columns = (2, 3, 15)
    sh = wb.active
    data = [[x[m].value for m in columns] for x in sh]
    result = {}
    for i in data:
        if i[1] not in ["Получен", "Возврат компенсирован"]:
            continue
        if x := result.get(i[0], {}):
            if x["status"] == "vozvrat-predvaritelno-poluchen":
                continue
            x["status"] = "vozvrat-predvaritelno-poluchen"
            x["comment"] = str(i[2])
            result[i[0]] = x
        else:
            x["status"] = (
                "vozvrat-predvaritelno-poluchen"
                if i[1] == "Получен"
                else "vozvrat-predvaritelno-kompensirovan"
            )
            x["comment"] = str(i[2])
            result[i[0]] = x
    return result


async def find_orders_in_crm(
    session: aiohttp.ClientSession,
    queue_batches: asyncio.Queue,
    queue_update_orders: asyncio.Queue,
    result_queue: asyncio.Queue,
):
    """searching for orders in crm, if order is not found it will go to result queue

    Args:
        session (aiohttp.ClientSession): async session
        queue_batches (asyncio.Queue): queue from which orders for searching will be taken
        queue_update_orders (asyncio.Queue): after searching orders will go here
        result_queue (asyncio.Queue): if the search did not find the order, it will go here
    """
    while True:
        batch = await queue_batches.get()
        url = "https://mag54.retailcrm.ru/api/v5/orders?limit=100&page=1"
        url += "&filter[externalIds][]=" + "&filter[externalIds][]=".join(
            [x for x in batch]
        )
        headers = {"X-API-KEY": CRM_API_KEY}
        async with session.get(url, headers=headers) as response:
            result = await response.json()

        result_polucheno_magazinom = [
            (x["externalId"], batch.pop(x["externalId"]))
            for x in result["orders"]
            if x["status"] in ["vozvrat-poluchen-magazinom", "vozvrat-poluchen-wb-fbo"]
        ]
        for order in result_polucheno_magazinom:
            to_result_data = [
                order[0],
                order[1]["status"],
                'Статус в CRM "Вовзрат получен магазином"',
            ]
            result_queue.put_nowait(to_result_data)

        update_orders = [
            (x["externalId"], batch.pop(x["externalId"]))
            for x in result["orders"]
            if x["status"]
            not in ["vozvrat-poluchen-magazinom", "vozvrat-poluchen-wb-fbo"]
        ]

        for order in update_orders:
            queue_update_orders.put_nowait(order)

        for order in batch:
            to_result_data = [order, batch[order]["status"], "заказ не найден в CRM"]
            result_queue.put_nowait(to_result_data)

        queue_batches.task_done()


async def order_change_status_crm(
    session: aiohttp.ClientSession,
    queue_update_orders: asyncio.Queue,
    result_queue: asyncio.Queue,
):
    """Change data in crm

    Args:
        session (aiohttp.ClientSession): session
        queue_update_orders (asyncio.Queue): queue from which orders for updating will be taken
        result_queue (asyncio.Queue): queue where result will be going
    """
    while True:
        order_id, order_data = await queue_update_orders.get()

        headers = {"X-API-KEY": CRM_API_KEY}
        url = f"https://mag54.retailcrm.ru/api/v5/orders/{order_id}/edit?by=externalId"
        order = {
            "status": order_data["status"],
            "managerComment": order_data["comment"],
        }
        request_body = {"order": json.dumps(order)}

        async with session.post(
            url=url, headers=headers, json=request_body
        ) as response:
            result = await response.json()

        if not result["success"]:
            order_data["status"] = "Не удалось изменить статус"
        to_result_data = [order_id] + list(order_data.values())
        result_queue.put_nowait(to_result_data)

        queue_update_orders.task_done()
        await asyncio.sleep(0.9)


def create_report(result_queue: asyncio.Queue) -> openpyxl.Workbook:
    """creating excel report from async queue

    Args:
        result_queue (asyncio.Queue): result queue

    Returns:
        openpyxl.Workbook: workbook
    """
    result_data = [result_queue.get_nowait() for x in range(result_queue.qsize())]
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in result_data:
        ws.append(row)
    return wb


async def process_return_xlsx(wb: openpyxl.Workbook, extractor: function):
    """function to manage asynchronous functionality

    Args:
        wb (openpyxl.Workbook): excel workbook
        extractor (function): function wich extract data from excel
    """
    data = extractor(wb)
    data = dict_batches_100(data)

    batch_queues = asyncio.Queue()
    update_orders_queue = asyncio.Queue()
    result_queue = asyncio.Queue()

    for i in data:
        batch_queues.put_nowait(i)

    async with aiohttp.ClientSession() as session:
        tasks = [
            asyncio.create_task(
                find_orders_in_crm(
                    session, batch_queues, update_orders_queue, result_queue
                )
            )
            for x in range(8)
        ]
        await batch_queues.join()
        for task in tasks:
            task.cancel()

        tasks = [
            asyncio.create_task(
                order_change_status_crm(session, update_orders_queue, result_queue)
            )
            for x in range(8)
        ]
        await update_orders_queue.join()
        for task in tasks:
            task.cancel()

    report_wb = create_report(result_queue)
    report_wb.save("report.xlsx")


EXTRACTOR_OPTIONS = {
    "ozon": extract_data_ozon,
    "wb": extract_data_wb,
    "yandex": extract_data_ya,
}


async def main(wb: openpyxl.Workbook, marketplace: str):
    """entry point for updating orders in crm from marketplace reports

    Args:
        wb (openpyxl.Workbook): excel workbook
        marketplace (str): marketplace key from EXTRACTOR_OPTIONS
    """
    await process_return_xlsx(wb, EXTRACTOR_OPTIONS[marketplace])


async def main_ya():
    wb_yandex = openpyxl.load_workbook("yandex_return.xlsx")
    await main(wb_yandex, "yandex")


async def main_ozon():
    wb_ozon = openpyxl.load_workbook("ozon_2.xlsx")
    await main(wb_ozon, "ozon")


async def main_wb():
    wb_workbook = openpyxl.load_workbook("return_wb.xlsx")
    await main(wb_workbook, "wb")


if __name__ == "__main__":
    asyncio.run(main_wb())
